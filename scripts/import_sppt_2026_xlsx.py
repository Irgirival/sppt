import json
import re
import urllib.parse
import urllib.request
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = Path(r'C:/Users/irgir/Downloads/SPPT PBB 2026 1.xlsx')
INDEX_PATH = ROOT / 'index.html'
TAHUN_PAJAK = 2026
SHEET_NAME = 'P 2026'
SUPABASE_URL = 'https://yhcizybaxhtohgmtlgpj.supabase.co'


def get_anon_key():
    text = INDEX_PATH.read_text(encoding='utf-8')
    m = re.search(r"const\s+SUPABASE_ANON\s*=\s*'([^']+)'", text)
    if not m or '...' in m.group(1):
        raise RuntimeError('SUPABASE_ANON key not found or redacted in index.html')
    return m.group(1)


def norm_nop(v):
    if v is None:
        return ''
    s = str(v).strip().replace(' ', '')
    s = re.sub(r'2026$', '', s).strip()
    # Keep canonical workbook dash form: 32.15-031.003-001.0001.0
    if '-' not in s and re.fullmatch(r'\d{18}', s):
        s = f'{s[0:2]}.{s[2:4]}-{s[4:7]}.{s[7:10]}-{s[10:13]}.{s[13:17]}.{s[17]}'
    return s


def parse_amount(v):
    if v is None or v == '':
        return 0
    return int(round(float(v)))


def parse_rt_rw(v):
    if v is None or v == '':
        return None, None, None
    raw = str(v).strip().upper()
    if raw == 'LUAR':
        return None, 'LUAR', raw
    digits = re.sub(r'\D', '', raw)
    if len(digits) == 1:
        digits = '0' + digits
    if len(digits) >= 2:
        return digits[:-1].zfill(2), digits[-1:].zfill(2), raw
    return None, raw, raw


def iso(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def load_records():
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    records = []
    for excel_row, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        nop = norm_nop(row[0] if len(row) > 0 else None)
        nama = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ''
        if not nop or '32.15' not in nop or not nama:
            continue
        rt, rw, raw_rt_rw = parse_rt_rw(row[2] if len(row) > 2 else None)
        tgl_bayar = iso(row[8] if len(row) > 8 else None)
        ket = str(row[10]).strip().upper() if len(row) > 10 and row[10] else ''
        pokok = parse_amount(row[3] if len(row) > 3 else 0)
        wajib = parse_amount(row[4] if len(row) > 4 else 0)
        is_lunas = ket == 'BAYAR' or bool(tgl_bayar)
        jumlah_dibayar = wajib if is_lunas else 0
        records.append({
            'nop': nop,
            'tahun_pajak': TAHUN_PAJAK,
            'nama_wp': nama,
            'rt': rt,
            'rw': rw,
            'alamat': None,
            'luas_tanah': None,
            'luas_bangunan': None,
            'pokok_pajak': wajib,
            'status_bayar': 'lunas' if is_lunas else 'belum',
            'tgl_bayar': tgl_bayar,
            'jumlah_dibayar': jumlah_dibayar,
            'metode_bayar': (str(row[9]).strip() if len(row) > 9 and row[9] else ('BAPENDA/SPPT 2026 XLSX' if is_lunas else None)),
            'denda': 0,
            'tahun_tunggakan': 0 if is_lunas else 1,
            'raw_excel_row': excel_row,
            'raw_sppt_ketetapan': 1,
            'raw_jumlah_ketetapan': wajib,
            'raw_sppt_kol5': None,
            'raw_jumlah_kol6': None,
            'raw_sppt_kol7': None,
            'raw_jumlah_kol8': None,
            'raw_sppt_realisasi': 1 if is_lunas else 0,
            'raw_jumlah_realisasi': jumlah_dibayar,
            'raw_persen_realisasi': None,
            'raw_sppt_tunggakan': 0 if is_lunas else 1,
            'raw_jumlah_tunggakan': 0 if is_lunas else wajib,
            'raw_persen_tunggakan': None,
        })
    return records


def request_json(method, path, key, data=None, headers=None):
    body = None if data is None else json.dumps(data).encode('utf-8')
    req = urllib.request.Request(
        SUPABASE_URL + path,
        data=body,
        method=method,
        headers={
            'apikey': key,
            'Authorization': 'Bearer ' + key,
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            **(headers or {}),
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            txt = resp.read().decode('utf-8')
            return resp.status, json.loads(txt) if txt else None, dict(resp.headers)
    except urllib.error.HTTPError as e:
        txt = e.read().decode('utf-8', errors='replace')
        raise RuntimeError(f'HTTP {e.code} {method} {path}: {txt[:1000]}')


def table_count(key, tahun):
    path = f'/rest/v1/sppt?select=id& tahun_pajak=eq.{tahun}'
    path = path.replace('?select=id& ', '?select=id&')
    status, data, headers = request_json('GET', path, key, headers={'Prefer': 'count=exact'})
    cr = headers.get('Content-Range', '')
    if '/' in cr:
        return int(cr.rsplit('/', 1)[1])
    return len(data or [])


def upsert_records(key, records, chunk=500):
    path = '/rest/v1/sppt?on_conflict=nop,tahun_pajak'
    done = 0
    for i in range(0, len(records), chunk):
        part = records[i:i+chunk]
        request_json('POST', path, key, part, headers={
            'Prefer': 'resolution=merge-duplicates,return=minimal'
        })
        done += len(part)
        print(f'upserted {done}/{len(records)}')


def fetch_all_sppt(key):
    cols = 'nop,nama_wp,rt,rw,pokok_pajak,status_bayar,jumlah_dibayar,tgl_bayar,metode_bayar'
    rows = []
    page_size = 1000
    start = 0
    while True:
        end = start + page_size - 1
        path = f'/rest/v1/sppt?select={urllib.parse.quote(cols)}&tahun_pajak=eq.{TAHUN_PAJAK}&order=raw_excel_row.asc'
        _, data, _ = request_json('GET', path, key, headers={'Range-Unit': 'items', 'Range': f'{start}-{end}'})
        data = data or []
        rows.extend(data)
        if len(data) < page_size:
            break
        start += page_size
    return rows


def verify(key):
    rows = fetch_all_sppt(key)
    return {
        'rows': len(rows),
        'distinct_nop': len({r['nop'] for r in rows}),
        'lunas': sum(1 for r in rows if r.get('status_bayar') == 'lunas'),
        'belum': sum(1 for r in rows if r.get('status_bayar') == 'belum'),
        'target': sum(int(r.get('pokok_pajak') or 0) for r in rows),
        'realisasi': sum(int(r.get('jumlah_dibayar') or 0) for r in rows),
        'rt_filled': sum(1 for r in rows if r.get('rt')),
        'rw_filled': sum(1 for r in rows if r.get('rw')),
        'rw_values': sorted({str(r.get('rw')) for r in rows if r.get('rw')})[:20],
        'first3': rows[:3],
    }


def main():
    key = get_anon_key()
    records = load_records()
    summary = {
        'xlsx_rows': len(records),
        'xlsx_distinct_nop': len({r['nop'] for r in records}),
        'xlsx_lunas': sum(1 for r in records if r['status_bayar'] == 'lunas'),
        'xlsx_belum': sum(1 for r in records if r['status_bayar'] == 'belum'),
        'xlsx_target': sum(r['pokok_pajak'] for r in records),
        'xlsx_realisasi': sum(r['jumlah_dibayar'] for r in records),
        'xlsx_rt_filled': sum(1 for r in records if r['rt']),
        'xlsx_rw_filled': sum(1 for r in records if r['rw']),
    }
    print(json.dumps({'source_summary': summary}, ensure_ascii=False, indent=2))
    upsert_records(key, records)
    print(json.dumps({'db_verification': verify(key)}, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
